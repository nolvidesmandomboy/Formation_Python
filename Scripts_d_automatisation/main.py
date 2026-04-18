#Fichier Texte
import os.path
import json
import sqlite3
import openpyxl
import smtplib #<- bibliothèque qui permet d'envoyer des emails avec pythons
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import requests

#Sur Python, on peut ouvrir, éditer, sauvegarder et fermer un fichier.

#f = open("Mon_fichier.txt", "w")
'''ici, le open permet d'ouvrir le fichier et à l'intérieur des parenthèses on a le nom du ficher et le mode (ici "w" signifie write donc on va écrire mais ça aurait pu être autre chose. Vu que c'est w, si le fichier existe ça va supprimer le contenu et nous donner la possibilité d'écrire un nouveau truc, au contraire si ça n'existe pas ça crée le nouveau fichier pour nous permettre d'écrire.)

Il faut faire aussi attention à ce que dans le terminal le chement soit le même que le fichier dans lequel on va éxécuter le script, c'est là qu'il va être enregistrer.
'''
#f.write("Salut c'est Ninho ! X)")

#f.close()

#Le fichier write doit bien être mis avant le close, c'est pas possible de modifier le fichier une fois que c'est fermer.

#Exercice 1

'''l = []

for i in range (0,10) :
    l.append(f"{str(i+1)}\n")

print (l)
    
fichier = open ("nombres.txt", "w")

fichier.writelines (l)

fichier.close()'''

#Lire un fichier texte
'''f = open("Mon_fichier.txt", "r")

texte = f.read() #<- permet de lire l'intégralité du fichier, on peut aussi spécifier le nombre de caractères à lire dans le fichier directement dans les parenthèses 
print(texte) 

f.close()'''

#Gestion des erreurs 
'''try :
    f = open("Mon_fichierss.txt", "r")
except FileNotFoundError :
    print("ERREUR, le fichier n'a pas été trouvé")
else:
    lecture = f.read()
    f.close()

filenam = "Mon_fichier.txt"
if os.path.exists(filenam):
    print ("le fichier existe")
else :
    print("le fichier n'existe pas")'''

#Mettre automatiquement le bon "/" dans le chemin des fichier, suppression et ajout de dossier 
'''filname = os.path.join ("dossier1", "Mon_fichier.txt") # vu que le slash est différent sur windows, mac ou Linux, le os.path.join permets de faire ça automatiquement
print ("Filename : ", filname)
os.path.mkdir("") #<- permet de créer un dossier 
os.path.rmdiv("") #<- prmet de supprimer un dossier'''

#Convertir un texte en JSON (sérialiser)
'''personne1 = {"age":30,
             "prenom":"Lucas",
             "travail":"géologue", 
             "patrimoine":100000}

personne_json = json.dumps(personne1) # Permet de convertir un dictionnaire en chaîne de caractères JSON.
j = open("fichier_prsonne_json","w")
j.writelines(personne_json)
j.close()'''

#convertir un JSON en texte (dictionnaire)
'''e = open ("fichier_prsonne_json","r")
donnees_json = e.read()
personne = json.loads(donnees_json) #permet de convertir le JSON en chaîne de caractère (dictionnaire)
e.close()

print (personne["prenom"])'''

#Bases de données : langage SQL 

'''
Création de tables : exemples d'artiste et d'album

artiste
- nom
- artiste_id
- album_id

album
- album_id
- artiste_id
- titre
- date de sortie

Pour créer les tabls en question: 

CREATE TABLE artiste (
    nom VARCHAR, 
    artiste_id INTEGER NOT NULL PRIMARY KEY, 
    album_id INTEGER)

CREATE TABLE album (
    album_id INTEGER NOT NULL PRIMARY KEY, 
    artiste_id INTEGER REFERENCES artiste, 
    titre VARCHAR,
    annee_sortie INTEGER)

INSERT INTO artiste (nom) VALUES ("Michael Jackson")
INSERT INTO album (titre, annee_sortie) VALUES ("Thriller", 1982)
'''

#SQLite : création de la table

#Etapes :
'''
1. connexion : "albums.db"
2. executer / curseur (c'est un objet intermédiaire pour accéder à la base de données)
3. commit pour s'assurer que tout est bien enregistrer 
4.fermer
''' 

'''connexion = sqlite3.connect("album2.db")

curseur = connexion.cursor() #<- c'est mieux de faire comme ça plutôt que de directement executer la requte sur la connexion car ça créera à chaqu fois un nouveau curseur, autant créer un seul qu'on va réutiliser.
#curseur.execute(requete_creer_tabla_artiste) 
curseur.execute("""
INSERT INTO artiste (nom) VALUES ("Michael Jackson")
""") #<- c'est ici qu'on mets toutes les requêtes
mj_id = curseur.lastrowid # <- permet de directement récuperer l'ID du curseur précédent (attention, change à chaque curseur)
curseur.execute('INSERT INTO album (artiste_id, titre, annee_sortie) VALUES (' + str(mj_id) + ', "Thriller", 1982)')
connexion.commit()
connexion.close()'''

#Lecture de base de données

'''connexion = sqlite3.connect("album2.db")

curseur = connexion.cursor() 
pnl = curseur.execute("""
SELECT titre, nom FROM artiste as ar JOIN album as al ON ar.nom = "PNL" AND al.artiste_id = 1 
""").fetchall()
print(pnl)
curseur.execute("""
SELECT * FROM artiste
""")
# listedartistes = curseur.fetchall() #<- permet de récuperer le résultat dans une variable, par contre faut que la base de données soit dans le même dossier que le code
# print (listedartistes)
connexion.commit()
connexion.close()'''

#Fichiers PDF 
"""
Avec la bibliothèque PyPDF2 on peut : 

- Combiner des PDF 
- superposer des pages, faire une rotation
- On ne peut pas rajouter du texte ou des images
"""
#Combiner des fichiers PDF 
#Ne marche pas, on passe au cours suivant

#Excel

'''wb = openpyxl.load_workbook("C:\\Formation_Python\\Scripts_d_automatisation\\octobre.xlsx")  #-> permet de charger le fichier excel
print(wb.sheetnames) # <- affiche les noms de tous les feuilles du fichier excel
sheet = wb['Feuil1'] # <- permett de récuperer une feuille du fichier
#cell = sheet["A1"] #<- récupère la celulle de la feuille du fichier
cell = sheet.cell(4,3) #<- permet d'accéder à des donnés du fichier excel grâce aux coordonnées lignes/colonnes
'''

#Lire les données de plusieurs fichiers Excel
"""wb1 = openpyxl.load_workbook("C:\\Formation_Python\\Scripts_d_automatisation\\octobre.xlsx",data_only=True) #<- le data only ici permet de récupérer uniquement les données même lorsque q'un calcul est effectué et non la formule. 
wb2 = openpyxl.load_workbook("C:\\Formation_Python\\Scripts_d_automatisation\\novembre.xlsx",data_only=True)
wb3 = openpyxl.load_workbook("C:\\Formation_Python\\Scripts_d_automatisation\\decembre.xlsx",data_only=True)
sheet1 = wb1['Feuil1']
def ajouter_data (wb, d):
    sheet = wb['Feuil1']
    for row in range (2,sheet.max_row):
        nom_articles = sheet.cell(row,1).value
        if not nom_articles:
            break
        print (nom_articles) #<- boucle pour récupérer toutes les valeurs d'une cellule (ici la liste des articles)
        total_ventes =  sheet.cell(row,4).value
        if d.get(nom_articles):
            d[nom_articles].append(total_ventes)
        else :
            d[nom_articles] = [total_ventes]

donnees = {}
ajouter_data(wb1,donnees)
ajouter_data(wb2,donnees)
ajouter_data(wb3,donnees)

print(donnees)"""

#Créer un nouveau fichier Excel

"""wb_sortie = openpyxl.Workbook() #permet de créer un nouveau fichier Excel
sheet_sortie = wb_sortie.active #permet de créer une feuille qui va être celle sur laquelle on va travailler 

def creer_colonne_ou_ligne (colonne_ou_row, valeur):
    for i in range (0, len(colonne_ou_row)) :
        sheet_sortie[colonne_ou_row[i]] = valeur[i]


colonnes = ["A1","B1","C1","D1"]
valeur1 = ["Articles","Octobre","Novembre","Décembre"]

# row = ["A2","A3","A4","A5"]
# valeur2 = ["Pommes","Poire","banane", "mangue","Ananas"]

creer_colonne_ou_ligne(colonnes,valeur1)
# creer_colonne_ou_ligne(row, valeur2)"""

#méthode de jonhathan
"""row=2
for i in donnees.items():
    nom_articles = i[0] #permet de récupérer les noms d'articles 
    ventes = i[1] #permet de récupérer les ventes
    sheet_sortie.cell(row,1).value = nom_articles #permet de mettre à jour les cellules
    for j in range (0,len(ventes)):
        sheet_sortie.cell(row,2+j).value = ventes[j]
    row+=1"""
#Créer un graphique 

"""chart_rf = openpyxl.chart.Reference(sheet_sortie, min_col =2,min_row=2,max_col=sheet_sortie.max_column,max_row=2) # Permet de donner ce qui va nous servir de référence dans nos données pour construire notre graphique
chart_serie = openpyxl.chart.Series(chart_rf,title="Total vente en Euro")
chart = openpyxl.chart.BarChart() #Permet de générer le graphique
chart.title = "Evolution du prix des pommes"
chart.append(chart_serie)

sheet_sortie.add_chart(chart,"F2") #Rajout le graphique dans notr feuille
wb_sortie.save("total_vente_trismestre.xlsx")"""

#Envoyer des mail 
"""config_email = "xxxx@gmail.com"
config_password = "xxxxx"
config_server = "smtp.gmail.com"
config_server_port = 587

def envoyer_mail (mail_destinataire,message, sujet):

    multipart_message = MIMEMultipart() #Avec ça, dans le mail on va pouvoir mettre plusieurs éléments (contenu texte, contenu HTML etc etc mais aussi confgurer d'autres éléments comme le sujet du message)
    multipart_message["Subject"] = sujet
    multipart_message["From"] = config_email
    multipart_message["To"] = mail_destinataire
    multipart_message.attach(MIMEText(message,  "plain")) #le Plain ici permet d'ajouter un message sous format texte, on pourrait le mettre sous format html plutôt par exemple

    serveur_mail = smtplib.SMTP (config_server, config_server_port)
    serveur_mail.starttls()
    serveur_mail.login(config_email,config_password)
    serveur_mail.sendmail(config_email,mail_destinataire,multipart_message.as_string())
    serveur_mail.quit()

message_email = '''
Bonjour bg je t'envoie ce message depuis python
'''

envoyer_mail("xxxx@gmail.com", message_email, "Email depuis python")"""

#Faire des appels réseaux

# https://codeavecjonathan.com/res/programmation.txt
# https://codeavecjonathan.com/res/pizzas1.json
# https://codeavecjonathan.com/res/exemple.html

reponse = requests.get("https://codeavecjonathan.com/res/programmation.txt")
reponse.encoding = "utf-8" #Permet de bien encoder le texte et gérer les accents par exemple 
print (reponse.text)

reponse2 = requests.get("https://codeavecjonathan.com/res/pizzas1.json")
print (reponse2.text)
pizza = json.loads(reponse2.text)
print (f"Nombre de pizzas : {len(pizza)}")